#!/usr/bin/env python3
"""
Gera o "fluxograma de exames" (planilha de UTI: exames nas linhas, datas nas colunas),
ORGANIZADO POR TEMA CLÍNICO, a partir do JSON de
/api/pacientes/:prontuario/exames?formato=resultados&incluirResultados=true

Uso:
    curl -s "http://localhost:3000/api/pacientes/574779/exames?formato=resultados&incluirResultados=true" -o /tmp/ex.json
    python3 scripts/fluxograma_exames.py /tmp/ex.json \
        --nome "DAVI OLIVEIRA DA CRUZ" --leito "UTI Neonatal — 2" \
        --out docs/fluxograma-leito2-574779.xlsx --md /tmp/fluxo.md

Requer: openpyxl  (pip install --break-system-packages openpyxl)

Notas (ver aprendizado 2026-06-16-fluxograma-exames-planilha-xlsx):
- O HICD NÃO retorna valores de referência (VR) neste formato.
- Diferencial leucocitário usa o valor relativo (_VR = %), não o absoluto (_VA).
- Todos os exames são agrupados por TEMA (Hemograma, Eletrólitos, Hepática, Lipidograma...).
- Sorologias (HIV/HBsAg/VDRL/Anti-HCV) NÃO vêm do endpoint de exames — vêm da evolução.
"""
import json, argparse, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# GRUPOS = [(tema, [(rótulo, sigla|None), ...])]
# sigla None = linha do modelo padrão que o HICD não fornece (mantida em branco).
GRUPOS = [
    ("HEMOGRAMA", [
        ("Hemoglobina", "HGB"), ("Hematócrito", "HTO"), ("Hemácias (RBC)", "RBC"),
        ("VCM", "VCM"), ("HCM", "HCM"), ("CHCM", "CHCM"), ("RDW", "RDW"),
        ("Leucócitos", "WBC"),
        ("Blastos %", "BLAST_VR"), ("Mielócitos %", "MIEL_VR"), ("Metamielócitos %", "META_VR"),
        ("Neutrófilos Bastões %", "BAST_VR"), ("Neutrófilos Segmentados %", "SEGM_VR"),
        ("Eosinófilos %", "EOSI_VR"), ("Basófilos %", "BASO_VR"),
        ("Linfócitos %", "LINF_VR"), ("Monócitos %", "MONO_VR"),
        ("Plaquetas", "PLT"),
    ]),
    ("COAGULAÇÃO", [
        ("TAP", "TAP_TEMPO_DO_P"), ("TTPA", "TTP"), ("RNI", "RNI"),
        ("Tempo de Sangramento", None),
    ]),
    ("ELETRÓLITOS", [
        ("Na+", "SOD"), ("K+", "POT"), ("CL-", "CLO"),
        ("Mg+2", "MAG"), ("Ca+2", "CAL"), ("Fósforo (P)", "FOS"),
        ("Na+ (íon/gaso)", "SOI"), ("K+ (íon/gaso)", "POS"), ("CL- (íon/gaso)", "CLR"),
    ]),
    ("FUNÇÃO RENAL", [
        ("Uréia", "URE"), ("Creatinina", "CRE"),
    ]),
    ("FUNÇÃO HEPÁTICA / PANCREÁTICA", [
        ("TGO", "TGO"), ("TGP", "TGP"), ("GGT", "GGT"), ("Fosfatase Alcalina", "FAL"),
        ("Bilirrubina Total", "BT"), ("Bilirrubina Direta", "BD"), ("Bilirrubina Indireta", "BI"),
        ("Amilase", "AMI"), ("DHL", "RESULTADO"),
    ]),
    ("PROTEÍNAS", [
        ("Proteínas Totais", "PROT_TOT"), ("Albumina", "ALBUMINA"), ("Globulina", "GLOB"),
    ]),
    ("LIPIDOGRAMA", [
        ("Colesterol Total", "COL"), ("LDL", "RESULTADO_LD"), ("HDL", "HDL"),
        ("VLDL", "COLESTEROL_V"), ("Triglicerídeos", "TRI"),
    ]),
    ("ENZIMAS MUSCULARES / CARDÍACAS", [
        ("CPK", "CPK"), ("CKMB", "CMB"),
    ]),
    ("MARCADORES INFLAMATÓRIOS / INFECÇÃO", [
        ("PCR", "PCR"), ("Procalcitonina", "PCA"), ("CRQ", "CRQ"),
    ]),
    ("NÍVEIS SÉRICOS (monitorização)", [
        ("Vancocinemia", "VAN"),
    ]),
    ("GASOMETRIA", [
        ("pH Arterial", "PH"), ("PO2", "PO2"), ("PCO2", "PCO2"),
        ("HCO3", "HCO3"), ("BE", "BE"), ("SaO2", "SAO2"), ("Lactato", "LAC"),
    ]),
    ("URINA", [
        ("Urina - Densidade", None), ("Urina - Uréia", None), ("Urina - Na+", None),
        ("Urina - K+", None), ("Urina - Cl-", None), ("Minibal", None), ("Urocultura", "URO"),
    ]),
    ("MICROBIOLOGIA", [
        ("Hemocultura", "HMC"),
    ]),
    ("IMAGEM", [
        ("RX de Tórax", "TRX"), ("RX de Abdome", "RAS"), ("RX (RAD)", "RAD"), ("ABP", "ABP"),
    ]),
    ("TIPAGEM SANGUÍNEA", [
        ("Grupo Sanguíneo (ABO)", "ABO"), ("Grupo Sanguíneo", "GCO"), ("Fator Rh", "FRH"),
    ]),
    ("SOROLOGIAS (origem: evolução, não o endpoint de exames)", [
        ("HIV", None), ("HbsAg", None), ("VDRL", None), ("Anti-HCV", None),
    ]),
]


def limpa(v):
    """Sanitiza valor para caber numa célula: 1 linha, sem boilerplate, extrai o essencial
    de blocos complexos (TTPA, tipagem) que o HICD devolve como texto cru."""
    if not v:
        return ""
    v = v.replace("\r", " ").replace("\n", " ")
    for marca in ("Liberado por", "Data do Cadastro", "Impressa", "X X X X"):
        i = v.find(marca)
        if i != -1:
            v = v[:i]
    if "EXAME NAO REAL" in v.upper():
        return "não realizado"
    if "Tempo do Paciente" in v:  # TTPA composto → tempo + INR
        m = re.search(r"Tempo do Paciente:?[->\s]*([\d.,]+)", v)
        r = re.search(r"\bR\.?[->\s]*([\d.,]+)", v)
        out = []
        if m: out.append(m.group(1) + "s")
        if r: out.append("INR " + r.group(1))
        if out: return " / ".join(out)
    if "Grupo Sanguineo" in v or "Fator Rh" in v:  # tipagem (casa só a linha com setas '--->')
        g = re.search(r"Grupo Sanguineo-{2,}>\s*([ABO]+)", v)
        rh = re.search(r"Fator Rh-{2,}>\s*([A-Z]+)", v)
        val = " ".join(x.group(1) for x in (g, rh) if x)
        if val:
            return val
    v = re.sub(r"[-=]{2,}>?", " ", v)        # runs de setas/pontos
    v = re.sub(r"\s+", " ", v).strip().replace("|", "/")
    return v[:48]


def carregar(path):
    j = json.loads(open(path, encoding="utf-8", errors="replace").read())
    # aceita tanto a resposta crua da API ("data") quanto o cache de exames_store ("bruto")
    fonte = j.get("bruto") if "bruto" in j else j.get("data", [])
    grade = {}
    for e in fonte:
        d = (e.get("data") or "")[:5]
        if not d:
            continue
        g = grade.setdefault(d, {})
        for r in (e.get("resultados") or []):
            s = r.get("sigla") or r.get("nome") or "?"
            v = limpa((r.get("resultado") or r.get("valor") or "").strip())
            if v and v != "0":
                g[s] = v
            elif s not in g:
                g[s] = v
    datas = sorted(grade.keys(), key=lambda d: (d.split("/")[1], d.split("/")[0]))
    return grade, datas


def gerar(path, nome, leito, out_xlsx, out_md=None):
    grade, datas = carregar(path)
    cell = lambda d, s: grade.get(d, {}).get(s, "") if s else ""

    if out_md:
        L = ["**NOME:** %s  **LEITO:** %s\n" % (nome, leito)]
        hdr = "| EXAMES | " + " | ".join(datas) + " |"
        sep = "|---|" + "|".join(["---"] * len(datas)) + "|"
        for tema, rows in GRUPOS:
            L.append("\n### %s\n" % tema)
            L += [hdr, sep]
            for rot, sig in rows:
                L.append("| " + rot + " | " + " | ".join(cell(d, sig) for d in datas) + " |")
        open(out_md, "w", encoding="utf-8").write("\n".join(L))

    wb = Workbook(); ws = wb.active; assert ws is not None; ws.title = "Exames"
    thin = Side(style="thin", color="999999"); border = Border(thin, thin, thin, thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    hl = PatternFill("solid", fgColor="EFEFEF"); grp = PatternFill("solid", fgColor="D9E1F2")
    ncol = 1 + len(datas)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, "HOSPITAL DE BASE — UTI NEONATAL — EXAMES").font = bold
    ws.cell(1, 1).alignment = center
    ws.cell(3, 1, "NOME: %s" % nome).font = bold
    ws.cell(3, 3, "LEITO: %s" % leito).font = bold
    H = 5
    ws.cell(H, 1, "EXAMES").font = bold; ws.cell(H, 1).fill = hl; ws.cell(H, 1).border = border
    for i, d in enumerate(datas):
        c = ws.cell(H, 2 + i, d); c.font = bold; c.alignment = center; c.fill = hl; c.border = border

    r = H + 1
    for tema, rows in GRUPOS:
        ct = ws.cell(r, 1, tema); ct.font = bold; ct.fill = grp; ct.border = border
        for i in range(len(datas)):
            ws.cell(r, 2 + i, "").fill = grp
        r += 1
        for rot, sig in rows:
            ws.cell(r, 1, rot).border = border
            for i, d in enumerate(datas):
                cc = ws.cell(r, 2 + i, cell(d, sig)); cc.alignment = center; cc.border = border
            r += 1

    ws.freeze_panes = "B%d" % (H + 1)
    ws.column_dimensions["A"].width = 26
    for i in range(len(datas)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 8
    wb.save(out_xlsx)
    return out_xlsx, len(datas)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("json")
    ap.add_argument("--nome", required=True)
    ap.add_argument("--leito", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--md", default=None)
    a = ap.parse_args()
    out, n = gerar(a.json, a.nome, a.leito, a.out, a.md)
    print("xlsx:", out, "| datas:", n, ("| md: " + a.md) if a.md else "")
